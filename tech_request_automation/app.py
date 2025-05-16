import os
import smtplib
import uuid
import logging
from logging.handlers import RotatingFileHandler
from email.mime.text import MIMEText
from flask import Flask, render_template, request, redirect, url_for, flash
import pandas as pd
from dotenv import load_dotenv
from datetime import datetime
import openpyxl
from openpyxl.utils import get_column_letter

# Load environment variables from .env file in the parent directory
load_dotenv(dotenv_path=os.path.join(os.path.dirname(__file__), '..', '.env'))

app = Flask(__name__, template_folder='../templates')

log_file_path = os.path.join(os.path.dirname(__file__), '..', 'app.log')
file_handler = RotatingFileHandler(log_file_path, maxBytes=10240, backupCount=10)
file_handler.setFormatter(logging.Formatter(
    '%(asctime)s %(levelname)s: %(message)s [in %(pathname)s:%(lineno)d]'
))
file_handler.setLevel(logging.INFO)
app.logger.addHandler(file_handler)
app.logger.setLevel(logging.INFO)

app.secret_key = os.urandom(24)

SMTP_SERVER = os.getenv('SMTP_SERVER', 'smtp.gmail.com')
SMTP_PORT = int(os.getenv('SMTP_PORT', 587))
SMTP_USERNAME = os.getenv('SMTP_USERNAME')
SMTP_PASSWORD = os.getenv('SMTP_PASSWORD')
IT_MANAGER_EMAIL = os.getenv('IT_MANAGER_EMAIL')
EXCEL_FILE = os.path.join(os.path.dirname(__file__), '..', 'requests.xlsx')

pending_requests = {}

def send_email(to_email, subject, body):
    try:
        msg = MIMEText(body, 'html')
        msg['Subject'] = subject
        msg['From'] = SMTP_USERNAME
        msg['To'] = to_email

        with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as server:
            server.starttls()
            server.login(SMTP_USERNAME, SMTP_PASSWORD)
            server.sendmail(SMTP_USERNAME, to_email, msg.as_string())
        return True
    except Exception as e:
        print(f"Error sending email: {e}")
        return False

@app.route('/', methods=['GET', 'POST'])
def submit_request_form():
    if request.method == 'POST':
        name = request.form['name']
        email = request.form['email']
        department = request.form['department']
        request_details = request.form['request_details']

        request_id = str(uuid.uuid4())
        pending_requests[request_id] = {
            'name': name,
            'email': email,
            'department': department,
            'request_details': request_details
        }

        approve_url = url_for('handle_approval', request_id=request_id, decision='approved', _external=True)
        reject_url = url_for('handle_approval', request_id=request_id, decision='rejected', _external=True)
        
        email_body = render_template('approval_email.html', 
                                     name=name, email=email, department=department, 
                                     request_details=request_details, 
                                     approve_url=approve_url, reject_url=reject_url)
        
        if send_email(IT_MANAGER_EMAIL, 'New Tech Solution Request', email_body):
            confirmation_body = render_template('confirmation_email.html', name=name, request_details=request_details)
            send_email(email, 'Request Confirmation', confirmation_body)
            flash('Request submitted successfully! An approval email has been sent to the IT Manager.', 'success')
        else:
            flash('Failed to send approval email. Please try again later.', 'error')
            if request_id in pending_requests:
                del pending_requests[request_id]
        
        return redirect(url_for('submit_request_form'))
    return render_template('form.html')

@app.route('/approval/<request_id>/<decision>')
def handle_approval(request_id, decision):
    app.logger.info(f"Handling approval for request_id: {request_id}, decision: {decision}")
    if request_id not in pending_requests:
        flash('Invalid or expired request link.', 'error')
        app.logger.warn(f"Invalid or expired request link used for request_id: {request_id}")
        return redirect(url_for('submit_request_form'))

    try:
        request_data = pending_requests.pop(request_id)
    except KeyError:
        flash('Request ID not found in pending requests. It might have been processed already.', 'error')
        app.logger.warn(f"Request ID {request_id} not found in pending_requests during pop.")
        return redirect(url_for('submit_request_form'))
        
    status = 'approved' if decision == 'approved' else 'rejected'
    app.logger.info(f"Request data for {request_id} popped. Status set to: {status}")

    if decision == 'approved':
        request_data['approval_date'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        try:
            df_new = pd.DataFrame([request_data])
            excel_dir = os.path.dirname(EXCEL_FILE)
            app.logger.info(f"Attempting to write to Excel file: {EXCEL_FILE}")
            app.logger.info(f"Target directory for Excel file: {excel_dir}")

            if not os.path.exists(excel_dir):
                app.logger.info(f"Excel directory {excel_dir} does not exist. This is unexpected.")

            if not os.access(excel_dir, os.W_OK):
                app.logger.error(f"Write permission DENIED for directory: {excel_dir}")
                flash('Critical Error: Permission denied to write to the Excel file directory.', 'error')
                pending_requests[request_id] = request_data
                return redirect(url_for('submit_request_form'))
            else:
                app.logger.info(f"Write permission GRANTED for directory: {excel_dir}")

            if os.path.exists(EXCEL_FILE):
                app.logger.info(f"Excel file found at {EXCEL_FILE}. Checking write permission for the file.")
                if not os.access(EXCEL_FILE, os.W_OK):
                    app.logger.error(f"Write permission DENIED for existing Excel file: {EXCEL_FILE}")
                    flash('Critical Error: Permission denied to write to the Excel file.', 'error')
                    pending_requests[request_id] = request_data
                    return redirect(url_for('submit_request_form'))
                else:
                    app.logger.info(f"Write permission GRANTED for existing Excel file: {EXCEL_FILE}")
                df_existing = pd.read_excel(EXCEL_FILE)
                for col in df_new.columns:
                    if col not in df_existing.columns:
                        df_existing[col] = pd.NA
                df_all = pd.concat([df_existing, df_new], ignore_index=True)
                app.logger.info(f"Data concatenated. Total rows now: {len(df_all)}")
            else:
                app.logger.info(f"Excel file not found at {EXCEL_FILE}. Creating new DataFrame.")
                df_all = df_new
            
            column_order = ['name', 'email', 'department', 'request_details', 'approval_date']
            df_all = df_all.reindex(columns=column_order)

            df_all.to_excel(EXCEL_FILE, index=False)
            app.logger.info(f"Successfully wrote to Excel file: {EXCEL_FILE}")

            try:
                workbook = openpyxl.load_workbook(EXCEL_FILE)
                worksheet = workbook.active
                for col_idx, column_cells in enumerate(worksheet.columns, 1):
                    max_length = 0
                    column_letter = get_column_letter(col_idx)
                    for cell in column_cells:
                        try:
                            if cell.value:
                                cell_length = len(str(cell.value))
                                if cell_length > max_length:
                                    max_length = cell_length
                        except:
                            pass
                    adjusted_width = (max_length + 2)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
                workbook.save(EXCEL_FILE)
                app.logger.info(f"Successfully auto-sized columns in Excel file: {EXCEL_FILE}")
            except Exception as e:
                app.logger.error(f"Error auto-sizing columns in Excel file {EXCEL_FILE}: {e}", exc_info=True)

            flash(f'Request {request_id} has been approved and logged.', 'success')
        except Exception as e:
            app.logger.error(f"CRITICAL ERROR during Excel operation for request {request_id} at path {EXCEL_FILE}: {e}", exc_info=True)
            flash(f'Error logging approved request. Please check application logs. Details: {str(e)[:100]}...', 'error')
            pending_requests[request_id] = request_data
            return redirect(url_for('submit_request_form'))
    else:
        app.logger.info(f"Request {request_id} was rejected. No Excel operation needed.")
        flash(f'Request {request_id} has been rejected.', 'success')

    app.logger.info(f"Preparing to send status update email for request {request_id}. Status: {status}")
    email_subject = f'Request {status.capitalize()}'
    status_update_body = render_template('status_update_email.html',
                                         name=request_data['name'],
                                         request_details=request_data['request_details'],
                                         status=status)
    if not send_email(request_data['email'], email_subject, status_update_body):
        app.logger.error(f"Failed to send status update email to {request_data['email']} for request {request_id}")
    
    return redirect(url_for('submit_request_form'))

if __name__ == '__main__':
    expected_columns = ['name', 'email', 'department', 'request_details', 'approval_date']
    if not os.path.exists(EXCEL_FILE):
        df_empty = pd.DataFrame(columns=expected_columns)
        df_empty.to_excel(EXCEL_FILE, index=False)
    else:
        try:
            df_check = pd.read_excel(EXCEL_FILE)
            missing_cols = [col for col in expected_columns if col not in df_check.columns]
            if missing_cols:
                app.logger.info(f"Existing Excel file is missing columns: {missing_cols}. Adding them.")
                for col in missing_cols:
                    df_check[col] = pd.NA
                df_check.to_excel(EXCEL_FILE, index=False)
        except Exception as e:
            app.logger.error(f"Error checking/updating columns in existing Excel file: {e}")

    app.run(debug=True)